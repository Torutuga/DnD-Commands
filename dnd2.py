from colorama import init
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as _gl
import os, random

os.system('cls')
init(autoreset=True)
book = load_workbook('C:/Users/hp/DnD/DnD.xlsx')
sheet3 = book["Map1"]

class Hex:
    def __init__(self, coor, bio, rel):
        self.coor = coor
        self.bio = bio
        self.rel = rel
        self.layer = False

def Color(v):
    par, inp = v.split('_')
    return '\033[' + par + 'm' + inp

def Zoning(sheet, c, v2):
    cl, rw = c.split(',')
    col = int(cl)
    row = int(rw)
    sheet[_gl(col) + str(row)] = str(v2)
    sheet[_gl(col-1) + str(row)] = str(v2)
    sheet[_gl(col-1) + str(row+1)] = str(v2)
    sheet[_gl(col) + str(row-1)] = str(v2)
    sheet[_gl(col) + str(row+1)] = str(v2)
    sheet[_gl(col) + str(row+2)] = str(v2)
    sheet[_gl(col+1) + str(row-1)] = str(v2)
    sheet[_gl(col+1) + str(row)] = str(v2)
    sheet[_gl(col+1) + str(row+1)] = str(v2)
    sheet[_gl(col+1) + str(row+2)] = str(v2)
    sheet[_gl(col+2) + str(row)] = str(v2)
    sheet[_gl(col+2) + str(row+1)] = str(v2)

def Size():
    x_cor = input('Ancho?\n...')
    y_cor = input('Alto?\n...')

    trueX = int(x_cor)*3 + 1
    trueY = int(y_cor)*4

    ## Extracts the Zones Coordinates
    zones = []
    for col in range(2, trueX, 3):
        for row in range(2, trueY, 4):
            if col%2 == 0:
                name = str(col) + ',' + str(row)
                zones.append(name)
            if col%2 != 0:
                n3 = row+2
                if n3 == trueY:
                    break
                else:
                    name = str(col) + ',' + str(n3)
                    zones.append(name)

    return trueX, trueY, zones

def Biomes(z):
    wtr = (len(z)*30)//100
    flt = (len(z)*65)//100
    pks = (len(z)*5)//100
    flt += len(z)-wtr-flt-pks
    swamp = flt//10  ## %10
    s_forest = (flt*2)//10  ## %20
    b_forest = flt//10  ## %10
    plains = (flt*4)//10  ## %40
    yards = (flt*2)//10  ## %20
    plains += flt-swamp-s_forest-b_forest-plains-yards
    snow = pks//10  ## %10
    mount = (pks*7)//10  ## %70
    b_stone = (pks*2)//10  ## %20
    mount += pks-mount-snow-b_stone

    needs = []
    needs.extend(['40;35;2_s' for x in range(swamp)])
    needs.extend(['40;32;2_t' for x in range(s_forest)])
    needs.extend(['40;32;2_T' for x in range(b_forest)])
    needs.extend(['43;30;2_"' for x in range(plains)])
    needs.extend(['43;30;2_n' for x in range(yards)])
    needs.extend(['40;30;1_^' for x in range(mount)])
    needs.extend(['47;30;1_~' for x in range(snow)])
    needs.extend(['40;31;2_n' for x in range(b_stone)])
    needs.extend(['40;36;2_~' for x in range(wtr)])

    while len(needs) != len(zones):
        if len(needs) > len(zones):
            needs.pop()
        if len(needs) < len(zones):
            needs.append('43;30;2_"')
        if len(needs) == len(zones):
            break

    return needs

def Printing(sht, zone, bio, y_cor, x_cor):
    for n in zone:
        indx = random.randint(0, len(bio)-1)
        zn = bio[indx]
        bio.remove(zn)
        Zoning(sht, n, zn)

    ## Fill Empty
    for y in range(1, y_cor+1):
        for x in range(1, x_cor+1):
            char = _gl(x)
            if not sht[char+str(y)].value:
                sht[char+str(y)] = '30;40;1_ ' 

def Relatives(x, y, zone):
    param = [(-3,-2), (0,-4), (3,-2), (-3,2), (0,4), (3,2)]
    rel = []
    for pair in param:
        place = str(int(x)+pair[0]) + ',' + str(int(y)+pair[1])
        if place in zone:
            rel.append(place)

    return rel

def Structure(sht, zone):
    data = []
    for coor in zone:
        x, y = coor.split(',')
        char = _gl(int(x))
        cell = sht[char+y].value
        rel = Relatives(x, y, zone)
        data.append(Hex(coor, cell, rel))

    return data

def Draw(sht, y_cor, x_cor):
    for y in range(1, y_cor+1):
        print('\n', end='')
        for x in range(1, x_cor+1):
            char = _gl(x)
            text = Color(sheet3[char+str(y)].value)
            print(text, end='')

col, row, zones = Size()
biomes = Biomes(zones)
Printing(sheet3, zones, biomes, row, col)
data = Structure(sheet3, zones)
Draw(sheet3, row, col)

print(len(zones))
response = input('\nSave?\n...')
if response == 'y':
    print('Alright!')
    book.save('C:/Users/hp/DnD/DnD.xlsx')
