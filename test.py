from colorama import init
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter as _gl
import os, random as ran, math

os.system('cls')
book = load_workbook('C:/Users/Hp/DnD/dnd.xlsx')
print(book.sheetnames)
sheet1 = book['Map1']

param = [(0,-1),(1,-1),
  (-1,0),(0, 0),(1, 0),(2,0),
  (-1,1),(0, 1),(1, 1),(2,1),
         (0, 2),(1, 2)]

class Zone:
    def __init__(self, cor, bio, rel):
        self.cor = cor
        self.bio = bio
        self.rel = rel
        self.limit = 6-len(rel)

def getCord(x, y):
    data = []
    for col in range(2, x+1, 3):
        for row in range(2, y+1, 4):
            if col%2 != 0:
                if row >= y-2:
                    pass
                else:
                    data.append((col, row+2))
            else:
                data.append((col, row))
    return data

def genBiomes(lst):
    data = []
    data.extend(['water' for x in range((len(lst)*3)//10)])
    data.extend(['peaks' for x in range(len(lst)//10)])
    data.extend(['flat' for x in range(len(lst)-len(data))])
    return data

def genRel(tpl, lst):
    data = []
    for p in lst:
        dst = math.sqrt(pow(tpl[0]-p[0], 2) + pow(tpl[1]-p[1], 2))
        if dst > 0 and dst <= 4:
            data.append(p)
        if len(data) >= 6:
            break
    return data

def genZones(lst1, lst2):
    data = []
    for z in lst1:
        bio = ran.randint(0, len(lst2)-1)
        rel = genRel(z, lst1)
        data.append(Zone(z, lst2[bio], rel))
        lst2.remove(lst2[bio])
    return data

def firstLayer(x, y):
    total_x = int(x)*3 +1
    total_y = int(y)*4
    zone_cor = getCord(total_x, total_y)
    biomes = genBiomes(zone_cor)
    return genZones(zone_cor, biomes), total_x, total_y

def riverRun(z):
    global center
    return math.sqrt(pow(z.cor[0]-center[0], 2) + pow(z.cor[1]-center[1], 2))

def Rivers(zones):

    data = []
    for z in zones:
        if z.bio == 'peaks':
            data.append(z)
    data.sort(reverse=True, key=riverRun)
    start = data.pop()

def Printing(sht):
    global zones, param

    bio = ''
    for z in zones:
        if z.bio == 'water':
            bio = '40;36;2_~'
        if z.bio == 'flat':
            bio = '43;30;2_"'
        if z.bio == 'peaks':
            bio = '40;30;1_^'
        for p in param:
            col = _gl(z.cor[0]+p[0])
            row = str(z.cor[1]+p[1])
            sht[col+row] = bio

def Color(v):
    if not v:
       return ' ' 
    else:
        par, inp = v.split('_')
        return '\033[' + par + 'm' + inp

def Draw(sht, y_cor, x_cor):
    for y in range(1, y_cor+1):
        print('\n', end='')
        for x in range(1, x_cor+1):
            char = _gl(x)
            text = Color(sht[char+str(y)].value)
            print(text, end='')

x = input('Ancho?\n...')
y = input('Alto?\n...')
zones, tx, ty = firstLayer(x, y)
center = (int(tx/2), int(ty/2))
Printing(sheet1)
init(autoreset=True)
Draw(sheet1, ty, tx)
#book.save('C:/Users/Hp/DnD/dnd.xlsx')
Rivers(zones)
