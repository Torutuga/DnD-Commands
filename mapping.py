from colorama import init
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as _gl
import os, random

os.system('cls')
init()
book = load_workbook('C:/Users/hp/DnD/DnD.xlsx')
sheet3 = book["Map1"]

def Color(v):
    par, inp = v.split('_')
    return '\033[' + par + 'm' + inp

def Zoning(sheet, c, v2):
    cl, rw = c.split(',')
    col = int(cl)
    row = int(rw)
    sheet[_gl(col) + str(row)] = '37;'+ str(v2) + ';22_0'
    sheet[_gl(col-1) + str(row)] = '37;' + str(v2) + ';22_0'
    sheet[_gl(col-1) + str(row+1)] = '37;' + str(v2) + ';22_0'
    sheet[_gl(col) + str(row-1)] = '37;' + str(v2) + ';22_0'
    sheet[_gl(col) + str(row+1)] = '37;' + str(v2) + ';22_0'
    sheet[_gl(col) + str(row+2)] = '37;' + str(v2) + ';22_0'
    sheet[_gl(col+1) + str(row-1)] = '37;' + str(v2) + ';22_0'
    sheet[_gl(col+1) + str(row)] = '37;' + str(v2) + ';22_0'
    sheet[_gl(col+1) + str(row+1)] = '37;' + str(v2) + ';22_0'
    sheet[_gl(col+1) + str(row+2)] = '37;' + str(v2) + ';22_0'
    sheet[_gl(col+2) + str(row)] = '37;' + str(v2) + ';22_0'
    sheet[_gl(col+2) + str(row+1)] = '37;' + str(v2) + ';22_0'

zones = []
for n1 in range(2, 17, 3):
    for n2 in range(2, 14, 4):
        if n1%2 == 0:
            name = str(n1) + ',' + str(n2)
            zones.append(name)
        if n1%2 != 0:
            n3 = n2+2
            if n2 == 10:
                break
            else:
                name = str(n1) + ',' + str(n3)
                zones.append(name)

needs = [0,0,0,0,1,1,1,1,1,1,1,2,2]
for n in zones:
    indx = random.randint(0, len(needs)-1)
    zn = needs[indx]
    needs.remove(zn)
    v2 = 0
    if zn == 0:
        v2 = '44'
    if zn == 1:
        v2 = '43'
    if zn == 2:
        v2 = '41'
    Zoning(sheet3, n, v2)

for row in range(1, 13):
    print('\n', end='')
    for col in range(1, 17):
        char = _gl(col)
        text = Color(sheet3[char+str(row)].value)
        print(text, end='')


print('\n', zones)


## Creating Kit
'''
for row in range(1, 13):
    print('\n', end='')
    for col in range(1, 17):
        char = _gl(col)
        sheet3[char+str(row)] = '37;40;22_N' 
'''
#book.save('C:/Users/hp/DnD/DnD.xlsx')
