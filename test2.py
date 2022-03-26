from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter as _gl
import math, random
from colorama import init

init(autoreset=True)
book = load_workbook('C:/Users/Hp/Dnd/dnd.xlsx')
mapsht = book['Map1']

zone = [31,20]

water = ((31*20)*4)/10
land = ((31*20)*6)/10

for x in range(1, 32):
    for y in range(1, 21):
        if water == 0:
            mapsht[_gl(x) + str(y)] = '43;30;2_"'
            land -=1
        if land == 0:
            mapsht[_gl(x) + str(y)] = '40;36;2_~'
            water -= 1
        else:
            result = random.randint(1,2)
            if result == 1:
                mapsht[_gl(x) + str(y)] = '43;30;2_"'
                land -= 1            
            if result == 2:
                mapsht[_gl(x) + str(y)] = '40;36;2_~'
                water -=1

def Color(v):
    if not v:
       return ' ' 
    else:
        par, inp = v.split('_')
        return '\033[' + par + 'm' + inp

def Draw(sht):
    for y in range(1, 21):
        print('\n', end='')
        for x in range(1, 30+1):
            char = _gl(x)
            text = Color(sht[char+str(y)].value)
            print(text, end='')

Draw(mapsht)
